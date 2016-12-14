# encoding:utf-8


import  re
import  os
import  locale
import xlrd
import  codecs
import  datetime
from openpyxl import Workbook, styles

import csv

nf = styles.numbers.FORMAT_NUMBER_COMMA_SEPARATED1
fn0 = styles.numbers.FORMAT_NUMBER_00


list_n_list = ['T', 'U']
fn0_list = ['M', 'N', 'O', 'P', "Q", "R"]

header_str_list=[
"Includes Amazon Marketplace, Fulfillment by Amazon (FBA), and Amazon Webstore transactions",																					
"All amounts in USD, unless specified",																						
"Definitions:",
"Sales tax collected: Includes sales tax collected from buyers for product sales, shipping, and gift wrap.",
"Selling fees: Includes variable closing fees and referral fees.",
"Other transaction fees: Includes shipping chargebacks, shipping holdbacks, per-item fees  and sales tax collection fees.",
'Other: Includes non-order transaction amounts. For more details, see the "Type" and "Description" columns for each order ID.'
]



def create_xls(**params):
    header = params.get("header", "")
    datas = params.get("datas", "")

    if not params.get("filename", ""):
        dd = datetime.datetime.now()
        filename = str(dd.strftime('%Y-%m-%d_%H-%M-%S')) + ".xlsx"
    else:
        filename = params.get("filename", "")
    print filename
    wb = Workbook()
    ws1 = wb.active
    ws1.title= "template"
    title_list = header
    for hsl in header_str_list:
        ws1.append(hsl)
    ws1.append(title_list)
    n = 1
    for lst in datas:
        tmp_list = []
        for i in lst:
            if lst.index(i)==0:
                tmp_list.append(datetime_to_str(i))
                pass
            else:
                if is_number(i):
                    tmp_list.append(deal_number(i))
                else:
                    tmp_list.append(i)
        ws1.append(list(tmp_list))
        # print lst[2].replace(" ", "").lower()
        if lst[2].replace(" ","").lower() == "transfer":
            for ni in list_n_list:
                _cell_ni = ws1.cell( ni + str(n + 1))
                _cell_ni.number_format = nf
        else:
           for lnl in fn0_list:
                _cell_lnl = ws1.cell( lnl + str(n))
                _cell_lnl.number_format = fn0
        n += 1
    return wb, filename


def create_csv(**params):
    header = params.get("header", "")
    datas = params.get("datas", "")
    print len(datas)
    if not params.get("filename", ""):
        dd = datetime.datetime.now()
        filename = str(dd.strftime('%Y-%m-%d_%H-%M-%S')) + ".xlsx"
    else:
        filename = params.get("filename", "")
    csvfile = open(filename,'wb')
    csvfile.write(codecs.BOM_UTF8)
    writer = csv.writer(csvfile)
    for hsl in header_str_list:
        header_hsl_list = [hsl]
        writer.writerow(header_hsl_list)
    writer.writerow(header)
    for lst in datas:
        tmp_list = []
        for i in lst:
            if lst.index(i) == 0:
                tmp_list.append(datetime_to_str(i))
            elif lst[2].replace(" ","").lower() == "transfer":
                tmp_list.append(csv_deal_number(i, tranfer=True))
            else:
                tmp_list.append(csv_deal_number(i, tranfer=False))
        try:
            writer.writerow(list(tmp_list))
        except Exception, e:
            pass
    csvfile.close()
    return filename



def is_number(strnum):
    """ 判断是否是正常的数字:或者是正常的小数"""
    # regex = re.compile(r"^(-?\d+)(\.\d*)?$")
    regex = re.compile(r"^[-+]{0,1}[0-9]{1,}.{0,1}[0-9]{0,}")
    if re.match(regex, strnum):
        return True
    else:
        return False


def deal_number(mnumber):
    """ 处理下金额数字，处理前面不带0,
        如, 0090 --> 90, 030.03 --> 30.03
    """
    try:
        af_money = int(mnumber)
    except:
        try:
            af_money = float(mnumber)
        except:
            af_money = mnumber
    return af_money

def csv_deal_number(mnumber, tranfer=None):
    if not is_number(mnumber):
        return mnumber
    else:
        try:
            mnumber = int(mnumber)
        except Exception, e:
            mnumber = mydeal_number(mnumber,tranfer=tranfer)
        return mnumber

def mydeal_number(numstr, tranfer):
    """ 如果小数点后面是0,则去掉"""
    numstr = str(numstr)
    numstr_list = numstr.split(".")
    if len(numstr_list) < 2:
        return numstr
    else:
        try:
            int(numstr_list[1])  #如果第二个不是数字
        except Exception, e:
            return numstr
        if int(numstr_list[1]) > 0:
            if tranfer:
                return round(float(numstr), 3)
            else:
                return round(float(numstr), 2)
        else:
            return numstr_list[0]


def datetime_to_str(dt):
    return dt.strftime("%b %d, %Y %I:%M:%S %p PDT") #PDT, PST







def generate_path(root_path):
    one_p_num = 1000
    two_p_num = 200
    i, j, k= 0, 0, 0
    while True:
        one_path = os.path.join(root_path, str(i))
        if not os.path.exists(one_path):
            os.mkdir(one_path)
        if len(os.listdir(one_path)) >= one_p_num:
            i += 1
            one_path = os.path.join(root_path, str(i))
            if not os.path.exists(one_path):
                os.mkdir(one_path)
            else:
                continue
        else:
            two_path = os.path.join(one_path, str(j))
            if not os.path.exists(two_path):
                os.mkdir(two_path)
            if len(os.listdir(two_path)) >= two_p_num:
                j+= 1
            else:
                three_path = os.path.join(two_path, str(k))
                if not os.path.exists(three_path):
                    os.mkdir(three_path)
                    return three_path
                else:
                    k += 1
