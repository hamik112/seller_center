#!/usr/bin/env python
# encoding:utf-8

import sys
import redis
import xlrd
import xlwt
import random
import os
reload(sys)
sys.setdefaultencoding("utf-8")

import  logging
from celery import  task
from celery.utils.log import get_task_logger
import datetime
import time
import csv
from openpyxl import load_workbook

from manageApp.models import  StatementView, FilenameToStorename,SkuProduct
from center.models import InventoryReportsData
from manageApp.dataService.tasks_util import update_file_statue, str_to_datetime
from manageApp.dataService.tasks_util import inventory_update_file_statue
from manageApp.dataService.deal_xls import read_xls, inventory_read_txt
from manageApp.dataService.csv_to_excel import csv_to_xls
from manageApp.dataService.statement_view_by_month import create_statement_month

from manageApp.handle_add_data import hanled as add_hanled
logger = get_task_logger(__name__)

log1 = logging.getLogger("tasks")


@task(max_retries=3,default_retry_delay=1 * 6)
def import_one_file_to_statement_view(file_path, filename):
    logger.info("running ...")
    print file_path, filename
    if filename.endswith(".csv"):
        file_path = csv_to_xls(file_path)
    try:
        datas = read_xls(file_path)
    except Exception, e:
        print str(e)
        datas = {"data": [],"statue": -1 ,"msg": "文件格式错误,请重新文件另外存为.xls文件再上传"}
    # filename = task_dict.get("filename", "")
    datas = datas.get("data", {})

    if isinstance(datas, dict) and  datas.get("statue", "") == -1 and datas.get("msg", ""):
        update_file_statue(filename, -1, error_msg=datas.get("msg", ""))
        return {"statue": -1, "msg": datas.get("msg", "")}

    value_list = []
    try:
        f_name = filename.split("__")[-1]
    except Exception, e:
        f_name = ""
        msg = "文件名错误!"
        update_file_statue(filename, -1, error_msg=msg)
        return {"statue": -1, "msg": msg}
    for dt in datas:
        if dt.get("name", "").lower() == "sheet1" or \
                        dt.get("name", "").lower() == "template" or \
                        dt.get("name","").replace(" ","") == f_name.split(".")[0].replace(" ", ""):
            value_list = dt.get("values", [])
            break
    if not value_list:
        try:
            value_list = datas[0].get("values", [])
        except Exception, e:
            msg = "没有找到文件!" + str(e)
            print msg
            log1.info(msg)
            update_file_statue(filename, -1, error_msg=msg)
    if not value_list:
        msg = "表格无数据或请查看是否在第一个sheet里面或查看sheet名称是sheet1"
        return {"statue": -1, "msg": msg}
        update_file_statue(filename, -1, error_msg=msg)
    # if len(value_list[3]) > 20:
    #     header_list = value_list[0]
    #     datas_list = value_list[1:]
    # else:
    header_list = value_list[7]
    datas_list = value_list[8:]
    need_header_list = ['date_time', 'settlement id', 'type', 'order id', 'sku', 'description', 'quantity',
                        'marketplace', 'fulfillment', 'order city', 'order state', 'order postal',
                        'product sales', "shipping credits", "gift wrap credits", "promotional rebates",
                        "sales tax collected", "selling fees", "fba fees", "other transaction fees",
                        "other", "total" ]
    header_dict = {}
    # print header_list,
    header_num_list = [0, 1, 2,3,4,5,6,7,8,9,10, 11,12,13,14,15,16,17,18,19,20,21]
    header_dict = dict(zip(need_header_list,header_num_list))
    """
    #实在没办法，根据头部字段来判断位置，他们写的不统一，实在难搞，直接上数字
    for name in need_header_list:
        try:
            if name == "date_time":
                header_dict[name] = header_list.index("date/time")
            else:
                header_dict[name] = header_list.index(name)
        except :
            if name == "date_time":
                msg = "没有找到字段: date/time"
            else:
                msg = "没有找到字段: " + str(name)
            update_file_statue(filename, -1, error_msg=msg)
            return {"statue": -1, "msg": msg}
    """
    # try:
    #     header_dict[u"店铺"] = header_list.index(u"店铺")
    # except Exception, e:
    #     msg = "没有找到字段: 店铺"
    #     update_file_statue(filename, -1, error_msg=msg)
    #     return {"statue": -1, "msg": msg}

    filename_split_list = filename.split("-")
    if len(filename_split_list) < 2:
        return {"statue": -1, "msg": "文件名错误:文件名格式不正确!"}
    serial_number = "-".join(filename.split("-")[:2])
    try:
        store_name = FilenameToStorename.objects.get(serial_number=serial_number).storename
    except Exception ,e:
        log1.info("filename: %s not found store_name"%str(filename))
        store_name = ""
    try:
        area  = filename.split(".")[0].split("-")[-1]
    except Exception, e:
        area = ""
    log1.info(len(value_list))
    n = 0
    stvs = []
    year = None
    month = None
    for data_line in datas_list:
        log1.info("current: "+ str(n))
        n += 1
        tmp_dict = {"filename": filename}
        for name in need_header_list:
            # if name == u"店铺" or name == "店铺":
            #     tmp_dict["store_name"] = data_line[header_dict.get(name)]
            if name == "date_time":
                print data_line[header_dict.get("date_time")]
                tmp_dict["date_time"] = str_to_datetime(data_line[header_dict.get("date_time")])
                year = str_to_datetime(data_line[header_dict.get("date_time")]).year
                month = str_to_datetime(data_line[header_dict.get("date_time")]).month
            else:
                dict_name = name.replace(" ", "_")
                print "name:", name
                if name == "total":
                    tmp_dict[dict_name] = str(data_line[header_dict.get(name)]).replace(',','')
                else:
                    tmp_dict[dict_name] = data_line[header_dict.get(name)]


        tmp_dict["serial_number"] = serial_number
        tmp_dict["store_name"] = store_name
        tmp_dict["area"] = area
        try:
            stv = StatementView(**tmp_dict)
            stvs.append(stv)
            stv.save()
        except Exception, e:
            print str(e)
            log1.info(str(e))
            try:
                StatementView.objects.filter(order_id=tmp_dict.get("order_id", "")).update(**tmp_dict)
            except Exception, e:
                msg = "存在重复记录, order_id是:"+str(tmp_dict.get("order_id", ""))
                log1.error(msg)
                update_file_statue(filename, -1, error_msg=msg)
                return {"statue": -1, "msg": str(e)}
    log1.info(str(n))
    update_file_statue(filename, 2)
    create_statement_month(serial_number,year,month,stvs,)
    return {"statue": 0, "msg": ""}


@task(max_retries=3,default_retry_delay=1 * 6)
def inventory_import(file_path, filename):
    print "filename:", filename
    print "file_path:",file_path
    if len(filename.split("__")) < 2:
        msg = u"文件名格式不对"
        inventory_update_file_statue(filename, -2, error_msg=msg)
        return {"status":-2, "msg":msg}

    username = filename.split("__")[0]    #所属用户，从文件名里面获取

    datas = inventory_read_txt(file_path)
    datas = datas.get("data", [])
    n = 0
    for line in datas:
        log1.info("current: "+ str(n))
        n += 1
        if len(line) < 6:
            continue
        try:
            tmp_dict = {"seller_sku": line[0], "fulfillment_channel_sku":line[1],
                    "asin":line[2],        "condition_type": line[3], 
                    "Warehouse_Condition_code":line[4], "Quantity_Available":line[5],
                    "filename":filename, "username":username
            }
            InventoryReportsData(**tmp_dict).save()
        except Exception, e:
            msg = "文件 %s 写入inveotry report错误: %s " % (filename,str(e))
            print msg
            log1.error(msg)
            inventory_update_file_statue(filename,-1, error_msg=msg)
    log1.info(str(n))
    inventory_update_file_statue(filename, 2)
    return {"status": 0, "msg": ""}

def read_csv(file_name):
    csvfile = open(file_name, 'rb')
    reader = csv.reader(csvfile,)
    return [row for row in reader]
def read_excel_xlsx(fil_name):
    wb = load_workbook(fil_name)
    sheet_names = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheet_names[0])
    rows_num = len(ws.rows)
    cols_num = len(ws.columns)
    datas = []
    for row in range(1,rows_num+1):
        data_items = []
        for col in range(1,cols_num+1):
            data_items.append(ws.cell(row=row, column=col).value)
        datas.append(data_items)
    return datas

def read_excel(file_name):
    """读取excel表格"""

    data = xlrd.open_workbook(file_name)
    sheet = data.sheets()[0]
    return  sheet._cell_values


@task(max_retries=3,default_retry_delay=1 * 6)
def deal_file(filepath,code,result_filepath,result_filepath2,obj,style):
    try:
        timenum = str(time.time()).replace('.','')
        r = redis.Redis(host='127.0.0.1', port='6379')
        file_type = filepath.split('.')[1]
        if file_type == 'xls':
            datas = read_excel(filepath)
        elif file_type == 'xlsx':
            datas = read_excel_xlsx(filepath)
        elif file_type == "csv":
            datas = read_csv(filepath)
        else:
            raise Exception('file type is bad')
        data_head = datas[:8]
        data_head[7].append('internal_sku')
        data_head[7].append('asin')
        data_head[7].append('hky')
        data_head[7].append('xuhao')
        data_head[7].append('wangguan')
        data_head[7].append('sku_price')
        data_head[7].append('sku_total_price')
        data_list = datas[8:]
        for li in data_list:
            """
            li[0]:date/time
            li[1]:settlement id
            li[2]:type
            li[3]:order id
            li[4]:sku
            li[5]:description
            li[6]:quantity
            li[7]:marketplace
            li[7]:fulfillment
            li[9]:order city
            li[10]:order state
            li[11]:order postal
            li[12]:product sales
            li[13]:shipping credits
            li[14]:gift wrap credits
            li[15]:promotional rebates
            li[16]:sales tax collected
            li[17]:selling fees
            li[18]:fba fees
            li[17]:other transaction fees
            li[20]:other
            li[21]:total
            """
            internal_sku = r.hget(li[4], 'sku')
            asin = r.hget(li[4], 'asin')
            li.append(internal_sku)
            li.append(asin)
            if internal_sku:
                if r.hget('asin_' + internal_sku, 'fba'):
                    li.append('haiyun')
                else:
                    li.append('kongyun')
            else:
                li.append('')
            try:
                sku_obj=SkuProduct.objects.filter(sku=internal_sku).first()
                sku_price=sku_obj.get_price()
            except:
                sku_price = ''
            if r.get('xh_'+code):
                li.append(code)
                li.append(r.get('xh_'+code))
                li.append(sku_price)
                try:
                    li.append(sku_price*int(li[6]))
                except:
                    li.append('')
            else:
                li.append('')
                li.append('')
                li.append(sku_price)
                try:
                    li.append(sku_price*int(li[6]))
                except:
                    li.append('')

        data_count = len(data_list)
        order_id_list_len = data_count * 3
        order_id_list = []
        for order_id in range(order_id_list_len):
            order_id_list.append(random.randint(1000000, 9999999))
        order_list = list(set(order_id_list))

        ss = 0
        address_list = []
        for li in data_list:
            if li[3]:
                address = [li[9], li[10], li[11]]
                address_list.append(address)
            else:
                continue
        random.shuffle(address_list)

        address_i = 0
        for li in data_list:
            """
            li[0]:date/time
            li[1]:settlement id
            li[2]:type
            li[3]:order id
            li[4]:sku
            li[5]:description
            li[6]:quantity
            li[7]:marketplace
            li[7]:fulfillment
            li[9]:order city
            li[10]:order state
            li[11]:order postal
            li[12]:product sales
            li[13]:shipping credits
            li[14]:gift wrap credits
            li[15]:promotional rebates
            li[16]:sales tax collected
            li[17]:selling fees
            li[18]:fba fees
            li[19]:other transaction fees
            li[20]:other
            li[21]:total
            li[22]:internal_sku
            li[23]:asin
            li[24]:海空运输
            li[25]:xuhao
            li[26]:wangguandaima
            """
            if li[3]:
                if li[2].lower() == 'chargeback refund' or li[2].lower() == 'a-to-z guarantee claim' or li[2].lower() == 'order' or li[2].lower() == 'refund':
                    if r.get(timenum+'addr_' + li[3]):
                        li[9] = eval(r.get(timenum+'addr_' + li[3]))[0]
                        li[10] = eval(r.get(timenum+'addr_' + li[3]))[1]
                        li[11] = eval(r.get(timenum+'addr_' + li[3]))[2]
                    else:
                        r.set(timenum+'addr_' + li[3], str(address_list[address_i]))
                        li[9] = address_list[address_i][0]
                        li[10] = address_list[address_i][1]
                        li[11] = address_list[address_i][2]
                if r.get(timenum+'orderid_'+li[3]):
                    li[3] = r.get(timenum+'orderid_'+li[3])
                else:
                    old_li3 = li[3]
                    li[3] = li[3][:4] + str(order_list[ss]) + '-'
                    ss += 1
                    li[3] += str(order_list[ss])
                    ss += 1
                    r.set(timenum+'orderid_'+old_li3, li[3])
                li[12] = (round(li[12] * 1.01, 2))
                li[21] = (
                float(li[12]) + float(li[13]) + float(li[14]) + float(li[15]) + float(li[16]) + float(li[17]) + float(
                    li[18]) + float(li[19]) + float(li[20]))
                if li[22]:
                    li[4] = r.hget('asin_' + li[23], 'newoutsku')
                if len(li) == 25:
                    print li[24]
                address_i += 1
            else:
                continue

        j = 0
        file = xlwt.Workbook()
        table = file.add_sheet('Sheet1', cell_overwrite_ok=True)
        for head in data_head:
            t = 0
            for h in head:
                table.write(j, t, head[t],style)
                t += 1
            j += 1

        for li in data_list:
            t = 0
            for h in li:
                table.write(j, t, li[t],style)
                t += 1
            j += 1
        os.popen("redis-cli KEYS '"+timenum+"addr_*' | xargs redis-cli DEL")
        os.popen("redis-cli KEYS '" + timenum + "orderid_*' | xargs redis-cli DEL")
        j = 0
        file.save(result_filepath)
        add_hanled(result_filepath,code,result_filepath2)
        obj.status = 1
        obj.save()
    except Exception,e:
        obj.status = 2
        obj.save()
        raise e